import openpyxl                 #Library for reading data from the excel sheet containing the data
import tkinter                  #Library for making dialog boxes
import math

wb = openpyxl.load_workbook("data.xlsx")                #Reading data for excel sheet
data = wb.active        
name=list()
tonset=list()
hod=list()
molwt=list()
hof=list()
heatcap=list()
for i in range(2, data.max_row+1):                              #Reading the required data and storing in lists        
        cobj = data.cell(row=i, column=1) 
        name.append(cobj.value)
for i in range(2, data.max_row+1):        
        cobj = data.cell(row=i, column=2)
        if cobj.value==None:
            continue
        else:
            tonset.append(float(cobj.value))
for i in range(2, data.max_row+1):       
        cobj = data.cell(row=i, column=3) 
        hod.append(float(cobj.value))
for i in range(2, data.max_row+1):       
        cobj = data.cell(row=i, column=4) 
        molwt.append(float(cobj.value))
for i in range(2, data.max_row+1):        
        cobj = data.cell(row=i, column=5)
        hof.append(float(cobj.value))
for i in range(2, data.max_row+1):        
        cobj = data.cell(row=i, column=6)
        heatcap.append(float(cobj.value))

name.insert(0,"Select")                 #Storing default values
tonset=[0]+tonset
hod=[0]+hod
molwt=[0]+molwt
hof=[0]+hof
heatcap=[0]+heatcap

for i in range(0,len(tonset)):
    if tonset[i]==None:
        tonset[i]=0

win=tkinter.Tk()                                        #Creating Dialog boxes
win.title("Reactants and Products input")
win.geometry("600x550")

root1=tkinter.Tk()
root1.title("Rate and Reaction Conditions Input")
root1.geometry("600x550")

root=tkinter.Tk()
root.title("User Input and Output")
root.geometry("600x500")


def index_subs(list_name,x):                            #Function to find the required properties of the substance selected
    idx=int(list_name.index(x))
    t1=tonset[idx]
    hd1=hod[idx]
    mwt=molwt[idx]
    hf1=hof[idx]
    hcap=heatcap[idx]
    return [t1,hd1,mwt,hf1,hcap]

def pen_tonset(x):                                      #Calculation of penalty factor for lowest temperature of decomposition
        if x>300:
                itonset=0
        elif x>200 and x<=300:
                itonset=1
        elif x>100 and x<=200:
                itonset=2
        elif x>50 and x<=100:
                itonset=3
        elif x<=50:
                itonset=4
        return itonset

def pen_mpd(x):                                         #Calculation of penalty factor for MPD
        if x<0.01:
                impd=0
        elif x>=0.01 and x<10:
                impd=1
        elif x>=10 and x<100:
                impd=2
        elif x>=100 and x<1000:
                impd=3
        elif x>=1000:
                impd=4
        return impd
def pen_del_tad(x):                                     #Calculation of penalty factor of adiabatic temperature rise
        if x<50:
                itad=1
        elif x>=50 and x<200:
                itad=2
        elif x>=200 and x<400:
                itad=3
        elif x>=400:
                itad=4
        return itad

def pen_heat(x):                                        #Calculation of penalty factor for enthalpy
        if x<100:
                iheat=1
        elif x>=100 and x<400:
                iheat=2
        elif x>=400 and x<800:
                iheat=3
        elif x>=800:
                iheat=4
        return iheat

def pen_tmrad(x):                                       #Calculation of penalty for TMRad
        if x>50:
                itmr=1
        elif x>24 and x<=50:
                itmr=2
        elif x>8 and x<=24:
                itmr=3
        elif x>1 and x<=8:
                itmr=4
        elif x<=1:
                itmr=5
        return itmr

def pen_cc(tp,mtsr,mtt,td24):                           #Calculation of penalty factor for criticality classes
        if tp<mtsr and mtsr<mtt and mtt<td24:
                icc=1
        elif tp<mtsr and mtsr<td24 and td24<mtt:
                icc=2
        elif tp<mtt and mtt<mtsr and mtsr<td24:
                icc=3
        elif tp<mtt and mtt<td24 and td24<mtsr:
                icc=4
        elif tp<td24 and td24<mtsr and mtsr<mtt:
                icc=5
        return icc
def ithiclass(x):
        if x<16:
                cl='I'
        elif x>=16 and x<32:
                cl='II'
        elif x>=32 and x<48:
                cl='III'
        elif x>=48 and x<64:
                cl='IV'
        elif x>=64:
                cl='V'
        return cl
def doh(x):
        if x=='I':
                haz='Light'
        elif x=='II':
                haz='Moderate'
        elif x=='III':
                haz='Intermediate'
        elif x=='IV':
                haz='Heavy'
        elif x=='V':
                haz='Severe'
        return haz

def sel_opt():
        res_label1.config(text="Reactant 1 selected: "+str(val_in1.get()))                                              #Displaying the user input
        res_st1.config(text="Stoichiometric coefficient of reactant 1 entered: "+ str(stcoeff1.get()))
        res_conc1.config(text="Concentration of reactant 1 entered: "+ str(conc1.get()))
        res_label2.config(text="Reactant 2 selected: "+str(val_in2.get()))
        res_st2.config(text="Stoichiometric coefficient of reactant 2 entered: "+ str(stcoeff2.get()))
        res_conc2.config(text="Concentration of reactant 2 entered: "+ str(conc2.get()))
        res_label3.config(text="Product 1 selected: "+str(val_in3.get()))
        res_st3.config(text="Stoichiometric coefficient of product 1 entered: "+ str(stcoeff3.get()))
        res_label4.config(text="Product 2 selected: "+str(val_in4.get()))
        res_st4.config(text="Stoichiometric coefficient of product 2 entered: "+ str(stcoeff4.get()))
                        
        r1=str(val_in1.get())                                                                                           #Storing values obtained from user in different variables
        [tonset1,hod1,molwt1,hof1,cp1]=index_subs(name,r1)
        r2=str(val_in2.get())
        [tonset2,hod2,molwt2,hof2,cp2]=index_subs(name,r2)
        p1=str(val_in3.get())
        [tonset3,hod3,molwt3,hof3,cp3]=index_subs(name,p1)
        p2=str(val_in4.get())
        [tonset4,hod4,molwt4,hof4,cp4]=index_subs(name,p2)

        st1=float(stcoeff1.get())
        st2=float(stcoeff2.get())
        st3=float(stcoeff3.get())
        st4=float(stcoeff4.get())

        c1=float(conc1.get())
        c2=float(conc2.get())
        
        
        m1=(c1*molwt1)/1000                     
        m2=(c2*molwt2)/1000
        cpmix=((m1*cp1)+(m2*cp2))/(m1+m2)                       #Calculating the reaction mixture heat capacity


        temp=(c1*st2)/st1

        hor=(st4*hof4)+(st3*hof3)
        hor=hor-((st2*hof2)+(st1*hof1))
        if c2<temp:                                             #Calculating the limiting reagent and setting it as the basis substance 
                MOLWT=molwt2
                hor=(hor*1000)/(st2*molwt2)
        else:
                MOLWT=molwt1
                hor=(hor*1000)/(st1*molwt1)
                
        if abs(hod1)>abs(hod2):
                HOD=abs(hod1)
                HOD1=(HOD*1000)/molwt1
                TONSET=tonset1
                del_tad_d=HOD1/cp1
                cp_final=cp1
        else:
                HOD=abs(hod2)
                HOD1=(HOD*1000)/molwt2
                TONSET=tonset2
                del_tad_d=HOD1/cp2
                cp_final=cp2

        del_tad_r=abs(hor)/cpmix
        
        # Incorporate thermal inertia into the adiabatic temperature rise calculation
        phi_value = float(phi.get())
        del_tad_r = abs(hor) / (cpmix * phi_value)
        del_tad_d = HOD1 / (cp_final * phi_value)


        return [HOD, TONSET, cpmix, hor, del_tad_d, del_tad_r, c1, c2, HOD1, cp_final]

def calc():
        res_tplabel.config(text="Process Temperature entered: "+str(tp.get()))
        res_mttlabel.config(text="MTT entered: "+str(mtt.get()))
        res_xrlabel.config(text="Degree of Accumulation entered: "+str(xr.get()))
        res_order1_label.config(text="Order wrt reactant 1 entered:"+str(order1.get()))
        res_order2_label.config(text="Order wrt reactant 2 entered:"+str(order2.get()))
        res_actenlabel.config(text="Activation energy of the reaction:"+str(acten.get()))
        res_alabel.config(text="Arrhenius factor:"+str(a.get()))
        res_td24_label.config(text="Td24 entered:"+str(td24.get()))
        res_order1_label.config(text="Order wrt reactant 1 entered:"+str(order1.get()))         #Displaying the values that the user entered
        res_order2_label.config(text="Order wrt reactant 2 entered:"+str(order2.get()))
        res_actenlabel.config(text="Activation energy of the reaction:"+str(acten.get()))
        res_alabel.config(text="Arrhenius factor:"+str(a.get()))
        res_td24_label.config(text="Td24 entered:"+str(td24.get()))
        [HOD,TONSET,cpmix,hor,del_tad_d,del_tad_r,c1,c2,HOD1,cp_final]=sel_opt()

        degacc=float(xr.get())
        Tp=float(tp.get())
        MTT=float(mtt.get())
        
        e=math.e
        A=float(a.get())
        A=A*(10**9)
        ea=float(acten.get())
        o1=float(order1.get())
        o2=float(order2.get())
        temp_1=(-1*ea*1000)/(8.3145*Tp)
        k=A*(e**temp_1)                                         #Calculating the rate constant maximum rate and 
        ratemax=k*(c1**o1)*(c2**o2)
        Td24=float(td24.get())

        mtsr=Tp+(degacc*del_tad_r)
        hor=hor+(cpmix*(Tp-298.15))
        HOD1=HOD1+(cp_final*(Tp-298.15))
        if hor<0:                                               #Checking if the reaction is exothermic
                if degacc>=0 and degacc<=1:
                        mpd=HOD*ratemax
                
                        TONSET=TONSET-273.15                            
            
                        penalty_tonset=pen_tonset(TONSET)               #Calculating the penalty                 
                        penalty_mpd=pen_mpd(mpd)
                        mf=1+((penalty_tonset*penalty_mpd)/16)
            
                        penalty_del_tad_reax=pen_del_tad(del_tad_r)
                        penalty_del_tad_decomp=pen_del_tad(del_tad_d)
                        penalty_hod=pen_heat(HOD1)
                        penalty_hor=pen_heat(hor)

                        sr=max(penalty_hor,penalty_del_tad_reax)
                        sdec=max(penalty_hod,penalty_del_tad_decomp)
                        s=sr+sdec

                        
                        # Adjust TMRad calculation with thermal inertia
                        qtd24 = (cpmix * 8.3145 * Td24 * Td24) / (24 * 3600 * ea * phi_value)
                        q = qtd24 * (math.e ** temp_1)
                        tmrad = (cpmix * 8.3145 * Tp * Tp) / (q * ea)
                        tmrad = tmrad / 3600
                        
                        penalty_tmrad=pen_tmrad(tmrad)
                        penalty_cc=pen_cc(Tp,mtsr,MTT,Td24)
                        p=penalty_tmrad+penalty_cc
                        ri=s*p                                          #Calculating the risk index and ithi
                        ithi=ri*mf
                        class_ithi=ithiclass(ithi)
                        hazard=doh(class_ithi)
                        result_label.config(text="ITHI value obtained:"+str(ithi))
                        final_label.config(text="The ITHI value falls in the range of Class "+class_ithi+". Hence, the degree of hazard is "+hazard)
                else:
                        result_label.config(text="Please enter a value for the degree of accumulation that is between 0 and 1!!!")
        else:
                result_label.config(text="Reaction is endothermic, hence runaway is not possible")

            

#Taking user input    
val_in1=tkinter.StringVar(win)
label1=tkinter.Label(win, text="Select Reactant 1:")
label1.pack()
val_in1.set("Select")
input_menu1=tkinter.OptionMenu(win,val_in1,*name)
input_menu1.pack()
stcoeff1=tkinter.StringVar(win)
stcoeff1_label = tkinter.Label(win, text = 'Enter the stoichiometric coefficient of reactant 1:')
stcoeff1_label.pack()
stcoeff1_entry = tkinter.Entry(win,textvariable = stcoeff1)
stcoeff1_entry.pack()
conc1=tkinter.StringVar(win)
conc1_label = tkinter.Label(win, text = 'Enter the concentration of reactant 1 mol/ml:')
conc1_label.pack()
conc1_entry = tkinter.Entry(win,textvariable = conc1)
conc1_entry.pack()

val_in2=tkinter.StringVar(win)
label2=tkinter.Label(win, text="Select Reactant 2:")
label2.pack()
val_in2.set("Select")
input_menu2=tkinter.OptionMenu(win,val_in2,*name)
input_menu2.pack()
stcoeff2=tkinter.StringVar(win)
stcoeff2_label = tkinter.Label(win,text='Enter the stoichiometric coefficient of reactant 2:')
stcoeff2_label.pack()
stcoeff2_entry = tkinter.Entry(win,textvariable = stcoeff2)
stcoeff2_entry.pack()
conc2=tkinter.StringVar(win)
conc2_label = tkinter.Label(win, text = 'Enter the concentration of reactant 2 in mol/ml:')
conc2_label.pack()
conc2_entry = tkinter.Entry(win,textvariable = conc2)
conc2_entry.pack()


val_in3=tkinter.StringVar(win)
label3=tkinter.Label(win, text="Select Product 1:")
label3.pack()
val_in3.set("Select")
input_menu3=tkinter.OptionMenu(win,val_in3,*name)
input_menu3.pack()
stcoeff3=tkinter.StringVar(win)
stcoeff3_label = tkinter.Label(win, text = 'Enter the stoichiometric coefficient of product 1:')
stcoeff3_label.pack()
stcoeff3_entry = tkinter.Entry(win,textvariable = stcoeff3)
stcoeff3_entry.pack()

val_in4=tkinter.StringVar(win)
label4=tkinter.Label(win, text="Select Product 2:")
label4.pack()
val_in4.set("Select")
input_menu4=tkinter.OptionMenu(win,val_in4,*name)
input_menu4.pack()
stcoeff4=tkinter.StringVar(win)
stcoeff4_label = tkinter.Label(win, text = 'Enter the stoichiometric coefficient of product 2:')
stcoeff4_label.pack()
stcoeff4_entry = tkinter.Entry(win,textvariable = stcoeff4)
stcoeff4_entry.pack()

tp=tkinter.StringVar(root1)
tp_label= tkinter.Label(root1,text='Enter the Process temperature in Kelvin:')
tp_label.pack()
tp_entry = tkinter.Entry(root1,textvariable=tp)
tp_entry.pack()

mtt=tkinter.StringVar(root1)
mtt_label= tkinter.Label(root1,text='Enter the MTT in Kelvin:')
mtt_label.pack()
mtt_entry = tkinter.Entry(root1,textvariable=mtt)
mtt_entry.pack()

xr=tkinter.StringVar(root1)
xr_label= tkinter.Label(root1,text='Enter the degree of accumulation:')
xr_label.pack()
xr_entry = tkinter.Entry(root1,textvariable=xr)
xr_entry.pack()

order1=tkinter.StringVar(root1)
order1_label= tkinter.Label(root1,text='Enter the order wrt reactant 1:')
order1_label.pack()
order1_entry = tkinter.Entry(root1,textvariable=order1)
order1_entry.pack()

order2=tkinter.StringVar(root1)
order2_label= tkinter.Label(root1,text='Enter the order wrt reactant 2:')
order2_label.pack()
order2_entry = tkinter.Entry(root1,textvariable=order2)
order2_entry.pack()

acten=tkinter.StringVar(root1)
acten_label= tkinter.Label(root1,text='Enter the activation energy in kJ/mol:')
acten_label.pack()
acten_entry = tkinter.Entry(root1,textvariable=acten)
acten_entry.pack()

a=tkinter.StringVar(root1)
a_label= tkinter.Label(root1,text='Enter the arrhenius factor in the factor of 10^9 s^-1:')
a_label.pack()
a_entry = tkinter.Entry(root1,textvariable=a)
a_entry.pack()

td24=tkinter.StringVar(root1)
td24_label= tkinter.Label(root1,text='Enter temperature at which tmrad=24 hrs in Kelvin:')
td24_label.pack()
td24_entry = tkinter.Entry(root1,textvariable=td24)
td24_entry.pack()

###new inertia code
phi = tkinter.StringVar(root1)
phi_label = tkinter.Label(root1, text='Enter the Thermal Inertia Factor (Phi):')
phi_label.pack()
phi_entry = tkinter.Entry(root1, textvariable=phi)
phi_entry.pack()

#Creating a submit button for dialog boxes
sub=tkinter.Button(win,text='Submit',command=sel_opt).pack()
sub=tkinter.Button(root1,text='Submit',command=calc).pack()

#Creating labels to display user input and the final result
res_label1=tkinter.Label(root,text="")
res_label1.pack()
res_st1=tkinter.Label(root,text="")
res_st1.pack()
res_conc1=tkinter.Label(root,text="")
res_conc1.pack()

res_label2=tkinter.Label(root,text="")
res_label2.pack()
res_st2=tkinter.Label(root,text="")
res_st2.pack()
res_conc2=tkinter.Label(root,text="")
res_conc2.pack()

res_label3=tkinter.Label(root,text="")
res_label3.pack()
res_st3=tkinter.Label(root,text="")
res_st3.pack()

res_label4=tkinter.Label(root,text="")
res_label4.pack()
res_st4=tkinter.Label(root,text="")
res_st4.pack()

res_tplabel=tkinter.Label(root,text="")
res_tplabel.pack()

res_mttlabel=tkinter.Label(root,text="")
res_mttlabel.pack()

res_xrlabel=tkinter.Label(root,text="")
res_xrlabel.pack()

res_order1_label=tkinter.Label(root,text="")
res_order1_label.pack()

res_order2_label=tkinter.Label(root,text="")
res_order2_label.pack()

res_actenlabel=tkinter.Label(root,text="")
res_actenlabel.pack()

res_alabel=tkinter.Label(root,text="")
res_alabel.pack()

res_td24_label=tkinter.Label(root,text="")
res_td24_label.pack()

result_label=tkinter.Label(root,text="")
result_label.pack()

final_label=tkinter.Label(root,text="")
final_label.pack()

win.mainloop()
root1.mainloop()
root.mainloop()

